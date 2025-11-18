"""
Export de DataFrames a Excel con manejo de tipos complejos.
"""

import io
import json
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def export_to_excel(
    df: pd.DataFrame,
    filename: str = "cv_analisis.xlsx",
    include_formatting: bool = True,
) -> io.BytesIO:
    """
    Exporta DataFrame a Excel con formato.

    Args:
        df: DataFrame con datos de CVs
        filename: Nombre del archivo (usado solo para referencia)
        include_formatting: Si incluir formato (colores, estilos)

    Returns:
        BytesIO con el archivo Excel
    """
    # Preparar datos para export
    df_export = _prepare_dataframe_for_export(df.copy())

    # Crear BytesIO buffer
    output = io.BytesIO()

    if include_formatting:
        # Export con formato usando openpyxl
        _export_with_formatting(df_export, output)
    else:
        # Export simple
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df_export.to_excel(writer, index=False, sheet_name="CVs")

    output.seek(0)
    return output


def _prepare_dataframe_for_export(df: pd.DataFrame) -> pd.DataFrame:
    """
    Prepara DataFrame para export, serializando tipos complejos.

    Args:
        df: DataFrame original

    Returns:
        DataFrame preparado para export
    """
    for col in df.columns:
        # Serializar listas y dicts a JSON string
        if df[col].dtype == "object":
            df[col] = df[col].apply(_serialize_value)

    return df


def _serialize_value(value: Any) -> str:
    """
    Serializa un valor complejo a string para Excel.

    Args:
        value: Valor a serializar

    Returns:
        String serializado
    """
    if value is None:
        return ""
    elif isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False, indent=None)
    elif isinstance(value, (int, float, bool)):
        return value
    else:
        return str(value)


def _export_with_formatting(df: pd.DataFrame, output: io.BytesIO):
    """
    Exporta con formato profesional.

    Args:
        df: DataFrame a exportar
        output: Buffer de salida
    """
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="CVs")

        workbook = writer.book
        worksheet = writer.sheets["CVs"]

        # Formato de headers
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True, size=11)

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Ajustar anchos de columna
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            # Limitar ancho máximo
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Freeze header row
        worksheet.freeze_panes = "A2"

        # Alinear texto en celdas
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=True
                )


def export_to_csv(df: pd.DataFrame) -> io.BytesIO:
    """
    Exporta DataFrame a CSV.

    Args:
        df: DataFrame con datos

    Returns:
        BytesIO con el CSV
    """
    df_export = _prepare_dataframe_for_export(df.copy())

    output = io.BytesIO()
    df_export.to_csv(output, index=False, encoding="utf-8")
    output.seek(0)

    return output


def export_to_json(df: pd.DataFrame) -> io.BytesIO:
    """
    Exporta DataFrame a JSON.

    Args:
        df: DataFrame con datos

    Returns:
        BytesIO con el JSON
    """
    output = io.BytesIO()

    # Convertir a JSON con formato legible
    json_str = df.to_json(orient="records", force_ascii=False, indent=2)
    output.write(json_str.encode("utf-8"))
    output.seek(0)

    return output


def create_summary_stats(df: pd.DataFrame) -> pd.DataFrame:
    """
    Crea estadísticas resumen del DataFrame de CVs.

    Args:
        df: DataFrame con datos de CVs

    Returns:
        DataFrame con estadísticas
    """
    stats = {
        "Total CVs": len(df),
        "CVs con error": len(df[df.get("error", "") != ""]),
        "CVs exitosos": len(df[df.get("error", "") == ""]),
    }

    # Agregar estadísticas específicas si existen columnas
    if "años_experiencia" in df.columns:
        valid_exp = pd.to_numeric(df["años_experiencia"], errors="coerce").dropna()
        if len(valid_exp) > 0:
            stats["Experiencia promedio (años)"] = round(valid_exp.mean(), 1)
            stats["Experiencia mínima (años)"] = int(valid_exp.min())
            stats["Experiencia máxima (años)"] = int(valid_exp.max())

    if "nivel_educativo_alcanzado" in df.columns:
        nivel_counts = df["nivel_educativo_alcanzado"].value_counts()
        stats["Nivel educativo más común"] = (
            nivel_counts.index[0] if len(nivel_counts) > 0 else "N/A"
        )

    # Convertir a DataFrame
    summary_df = pd.DataFrame([stats]).T
    summary_df.columns = ["Valor"]
    summary_df.index.name = "Métrica"

    return summary_df
